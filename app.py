from flask import Flask, render_template, request, send_file, jsonify
import os, json
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as ExcelImage
from datetime import datetime
import openai
from dotenv import load_dotenv
from parser.burp_parser import parse_burp_xml
from parser.semgrep.semgrep_scanner import run_semgrep_scan

load_dotenv()
openai.api_key = os.getenv("OPENAI_API_KEY")
app = Flask(__name__)

@app.route("/", methods=["GET", "POST"])
def index():
    results = []
    message = None
    severity_summary = {"High": 0, "Medium": 0, "Low": 0}

    if request.method == "POST":
        file = request.files.get("file")
        if file:
            filename = file.filename
            extension = os.path.splitext(filename)[1].lower()
            try:
                if extension == ".xml":
                    file.save("uploaded_report.xml")
                    results = parse_burp_xml("uploaded_report.xml")
                    message = "📄 تم تحليل تقرير Burp بنجاح!"
                elif extension in [".py", ".php", ".js", ".html"]:
                    os.makedirs("temp_code", exist_ok=True)
                    file_path = os.path.join("temp_code", filename)
                    file.save(file_path)
                    results = run_semgrep_scan("temp_code")
                    message = "🔍 تم تحليل الكود باستخدام Semgrep!"
                else:
                    message = "❌ نوع الملف غير مدعوم."
                with open("results.json", "w", encoding="utf-8") as f:
                    json.dump(results, f, ensure_ascii=False, indent=2)
                for r in results:
                    sev = r.get("severity", "").capitalize()
                    if sev in severity_summary:
                        severity_summary[sev] += 1
            except Exception as e:
                message = f"⚠️ خطأ أثناء التحليل: {str(e)}"
                results = []
    return render_template("index.html", results=results, message=message, summary=severity_summary)

@app.route("/download_excel", methods=["POST"])
def download_excel():
    with open("results.json", "r", encoding="utf-8") as f:
        results = json.load(f)
    if not results:
        return "❌ لا توجد بيانات حالية للتحويل."
    df = pd.DataFrame(results)
    wb = Workbook()
    ws = wb.active
    ws.title = "نتائج التحليل"
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws.merge_cells("A1:I1")
    ws["A1"] = f"تاريخ التحليل: {now}"
    ws["A1"].font = Font(bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="00C3FF")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_num, column_title in enumerate(df.columns, 1):
        cell = ws.cell(row=2, column=col_num, value=column_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        ws.column_dimensions[cell.column_letter].width = 40
    for r_idx, row in enumerate(df.itertuples(index=False), 3):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=str(value)).alignment = align_center
    if "severity" in df.columns:
        severity_counts = df["severity"].value_counts()
        plt.figure(figsize=(5, 5))
        plt.pie(severity_counts, labels=severity_counts.index, autopct="%1.1f%%", startangle=90)
        plt.title("📊 توزيع مستويات الخطورة", fontsize=12)
        plt.tight_layout()
        chart_path = os.path.join("static", "severity_chart.png")
        plt.savefig(chart_path, format="png")
        plt.close()
        if os.path.exists(chart_path):
            img = ExcelImage(chart_path)
            img.width = 300
            img.height = 300
            img.anchor = "J3"
            ws.add_image(img)
    output_path = "exports/analysis_results_one.xlsx"
    wb.save(output_path)
    return send_file(output_path, as_attachment=True)

@app.route("/download_excel_append", methods=["POST"])
def download_excel_append():
    with open("results.json", "r", encoding="utf-8") as f:
        results = json.load(f)
    if not results:
        return "❌ لا توجد بيانات حالية للتحويل."
    df = pd.DataFrame(results)
    file_path = "exports/analysis_results.xlsx"
    os.makedirs("exports", exist_ok=True)
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
        start_row = ws.max_row + 2
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "نتائج التحليل"
        ws.merge_cells("A1:I1")
        ws["A1"] = "📊 سجل تحاليل Burp Smart Eye"
        ws["A1"].font = Font(bold=True)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="00C3FF")
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws.cell(row=2, column=col_num, value=column_title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            ws.column_dimensions[cell.column_letter].width = 40
        start_row = 3
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    for i, row in df.iterrows():
        for j, value in enumerate(row, 1):
            ws.cell(row=start_row + i, column=j, value=str(value))
        ws.cell(row=start_row + i, column=len(df.columns) + 1, value=now)
    wb.save(file_path)
    return send_file(file_path, as_attachment=True)

@app.route("/ai_explain", methods=["POST"])
def ai_explain():
    vuln_name = request.form.get("vuln_name", "")
    if not vuln_name:
        return jsonify({"explanation": "❌ لا يوجد اسم للثغرة"})
    prompt = f"""اشرح الثغرة التالية ببساطة، ثم اذكر كيف يتم استغلالها وطريقة الحماية منها:\n{vuln_name}"""
    try:
        response = openai.ChatCompletion.create(
            model="gpt-4",
            messages=[
                {"role": "system", "content": "أنت خبير أمن سيبراني تشرح الثغرات ببساطة ودقة."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.5,
            max_tokens=500
        )
        explanation = response.choices[0].message.content
        return jsonify({"explanation": explanation})
    except Exception as e:
        return jsonify({"explanation": f"⚠️ خطأ أثناء توليد الشرح: {str(e)}"})

if __name__ == "__main__":
    app.run(debug=True)
